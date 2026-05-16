import pandas as pd
from openpyxl import load_workbook

# 1. Calculate the data
df = pd.read_excel("British Airways Summer Schedule Dataset - Forage Data Science Task 1.xlsx")
df['TOTAL_SEATS'] = df['FIRST_CLASS_SEATS'] + df['BUSINESS_CLASS_SEATS'] + df['ECONOMY_SEATS']

# Group by HAUL
haul_stats = df.groupby('HAUL').agg({
    'TIER1_ELIGIBLE_PAX': 'mean',
    'TIER2_ELIGIBLE_PAX': 'mean',
    'TIER3_ELIGIBLE_PAX': 'mean',
    'TOTAL_SEATS': 'mean'
})

haul_stats['T1_PCT'] = (haul_stats['TIER1_ELIGIBLE_PAX'] / haul_stats['TOTAL_SEATS'])
haul_stats['T2_PCT'] = (haul_stats['TIER2_ELIGIBLE_PAX'] / haul_stats['TOTAL_SEATS'])
haul_stats['T3_PCT'] = (haul_stats['TIER3_ELIGIBLE_PAX'] / haul_stats['TOTAL_SEATS'])

# Get example destinations for each haul
short_haul_dest = df[df['HAUL'] == 'SHORT']['ARRIVAL_COUNTRY'].unique()[:3].tolist()
long_haul_dest = df[df['HAUL'] == 'LONG']['ARRIVAL_COUNTRY'].unique()[:3].tolist()

# 2. Load the template
file_path = "Lounge Eligibility Lookup Template - Task 1.xlsx"
wb = load_workbook(file_path)
ws = wb["Lounge Eligibility Lookup Table"]

# Data to write (starting from row 11 based on previous inspection)
# Col 3: Grouping, Col 4: Examples, Col 5: T1%, Col 6: T2%, Col 7: T3%, Col 8: Notes

# Short Haul
ws.cell(row=11, column=4).value = "Short Haul (Europe/Turkey)"
ws.cell(row=11, column=5).value = ", ".join(short_haul_dest)
ws.cell(row=11, column=6).value = haul_stats.loc['SHORT', 'T1_PCT']
ws.cell(row=11, column=7).value = haul_stats.loc['SHORT', 'T2_PCT']
ws.cell(row=11, column=8).value = haul_stats.loc['SHORT', 'T3_PCT']
ws.cell(row=11, column=9).value = "Higher density of tier members per seat."

# Long Haul
ws.cell(row=12, column=4).value = "Long Haul (Global Hubs)"
ws.cell(row=12, column=5).value = ", ".join(long_haul_dest)
ws.cell(row=12, column=6).value = haul_stats.loc['LONG', 'T1_PCT']
ws.cell(row=12, column=7).value = haul_stats.loc['LONG', 'T2_PCT']
ws.cell(row=12, column=8).value = haul_stats.loc['LONG', 'T3_PCT']
ws.cell(row=12, column=9).value = "Lower density but higher total volume and premium seat count."

# Format percentages
for row in [11, 12]:
    for col in [6, 7, 8]:
        ws.cell(row=row, column=col).number_format = '0.00%'

# Populate Justification
ws_j = wb["Justification"]
ws_j.cell(row=12, column=4).value = "Flights were grouped by 'HAUL' type (Short vs Long) because this feature showed the most significant difference in the ratio of tier-eligible passengers to total seats available."
ws_j.cell(row=13, column=4).value = "Haul type captures both geographic region and aircraft configuration differences, which are primary drivers of customer loyalty distribution in this dataset."
ws_j.cell(row=14, column=4).value = "We assume that the 'eligible' counts provided in the dataset represent the average historical demand for lounge access per flight for those tiers."
ws_j.cell(row=15, column=4).value = "The model uses percentages of total capacity, making it scalable to different aircraft types or increased flight frequencies in the future."

wb.save(file_path)
print("Populated Lounge Eligibility Lookup Template successfully.")
